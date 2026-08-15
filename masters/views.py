import csv
import io
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse

from accounts.permissions import admin_required
from accounts.models import Role
from audit.services import log_action
from .models import Department, Programme, Specialization, HOD
from .forms import (
    DepartmentForm, ProgrammeForm, SpecializationForm, HODForm,
    DoctorForm, DoctorScheduleFormSet,
    StudentForm, StudentImportForm, StaffForm, StaffFamilyForm,
)
from doctors.models import Doctor
from patients.models import Patient, StudentProfile, StaffProfile, StaffFamilyProfile


# ── Hub ─────────────────────────────────────────────────────
@admin_required
def master_index(request):
    context = {
        'counts': {
            'departments': Department.objects.filter(status='Active').count(),
            'programmes': Programme.objects.filter(status='Active').count(),
            'specializations': Specialization.objects.filter(status='Active').count(),
            'hods': HOD.objects.filter(status='Active').count(),
            'doctors': Doctor.objects.filter(status='Active').count(),
            'students': Patient.objects.filter(category=Patient.CATEGORY_STUDENT, status='Active').count(),
            'staff': Patient.objects.filter(category=Patient.CATEGORY_STAFF, status='Active').count(),
            'family': Patient.objects.filter(category=Patient.CATEGORY_STAFF_FAMILY, status='Active').count(),
        }
    }
    return render(request, 'masters/index.html', context)


# ── Generic helpers ─────────────────────────────────────────
def _paginate(request, qs, per_page=25):
    return Paginator(qs, per_page).get_page(request.GET.get('page'))


# ═══════════════════════════════════════════════════════════
# DEPARTMENT
# ═══════════════════════════════════════════════════════════
@admin_required
def department_list(request):
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    qs = Department.objects.annotate(prog_count=Count('programmes')).order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(department_code__icontains=q))
    if status:
        qs = qs.filter(status=status)
    return render(request, 'masters/department_list.html', {
        'page_obj': _paginate(request, qs), 'q': q, 'status': status,
    })


@admin_required
def department_create(request):
    form = DepartmentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        log_action(request.user, 'create', 'masters', obj.pk, f'Department {obj}', request=request)
        messages.success(request, f'Department "{obj.name}" created.')
        return redirect('masters:department_list')
    return render(request, 'masters/department_form.html', {'form': form, 'title': 'Add Department'})


@admin_required
def department_edit(request, pk):
    obj = get_object_or_404(Department, pk=pk)
    form = DepartmentForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_action(request.user, 'update', 'masters', obj.pk, f'Department {obj}', request=request)
        messages.success(request, f'Department "{obj.name}" updated.')
        return redirect('masters:department_list')
    return render(request, 'masters/department_form.html', {'form': form, 'title': f'Edit Department – {obj.name}', 'object': obj})


@admin_required
def department_deactivate(request, pk):
    obj = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        obj.status = 'Inactive' if obj.status == 'Active' else 'Active'
        obj.save(update_fields=['status'])
        log_action(request.user, 'status_change', 'masters', obj.pk, f'Department {obj} → {obj.status}', request=request)
        messages.success(request, f'Department "{obj.name}" set to {obj.status}.')
    return redirect('masters:department_list')


# ═══════════════════════════════════════════════════════════
# PROGRAMME
# ═══════════════════════════════════════════════════════════
@admin_required
def programme_list(request):
    q = request.GET.get('q', '').strip()
    dept = request.GET.get('department', '')
    qs = Programme.objects.select_related('department').order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(programme_code__icontains=q))
    if dept:
        qs = qs.filter(department_id=dept)
    return render(request, 'masters/programme_list.html', {
        'page_obj': _paginate(request, qs), 'q': q, 'dept': dept,
        'departments': Department.objects.filter(status='Active'),
    })


@admin_required
def programme_create(request):
    form = ProgrammeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        log_action(request.user, 'create', 'masters', obj.pk, f'Programme {obj}', request=request)
        messages.success(request, f'Programme "{obj.name}" created.')
        return redirect('masters:programme_list')
    return render(request, 'masters/programme_form.html', {'form': form, 'title': 'Add Programme'})


@admin_required
def programme_edit(request, pk):
    obj = get_object_or_404(Programme, pk=pk)
    form = ProgrammeForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_action(request.user, 'update', 'masters', obj.pk, f'Programme {obj}', request=request)
        messages.success(request, f'Programme "{obj.name}" updated.')
        return redirect('masters:programme_list')
    return render(request, 'masters/programme_form.html', {'form': form, 'title': f'Edit Programme – {obj.name}', 'object': obj})


@admin_required
def programme_deactivate(request, pk):
    obj = get_object_or_404(Programme, pk=pk)
    if request.method == 'POST':
        obj.status = 'Inactive' if obj.status == 'Active' else 'Active'
        obj.save(update_fields=['status'])
        messages.success(request, f'Programme "{obj.name}" set to {obj.status}.')
    return redirect('masters:programme_list')


# ═══════════════════════════════════════════════════════════
# SPECIALIZATION
# ═══════════════════════════════════════════════════════════
@admin_required
def specialization_list(request):
    q = request.GET.get('q', '').strip()
    qs = Specialization.objects.annotate(doc_count=Count('doctors')).order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    return render(request, 'masters/specialization_list.html', {
        'page_obj': _paginate(request, qs), 'q': q,
    })


@admin_required
def specialization_create(request):
    form = SpecializationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        log_action(request.user, 'create', 'masters', obj.pk, f'Specialization {obj}', request=request)
        messages.success(request, f'Specialization "{obj.name}" created.')
        return redirect('masters:specialization_list')
    return render(request, 'masters/specialization_form.html', {'form': form, 'title': 'Add Specialization'})


@admin_required
def specialization_edit(request, pk):
    obj = get_object_or_404(Specialization, pk=pk)
    form = SpecializationForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f'Specialization "{obj.name}" updated.')
        return redirect('masters:specialization_list')
    return render(request, 'masters/specialization_form.html', {'form': form, 'title': f'Edit – {obj.name}', 'object': obj})


@admin_required
def specialization_delete(request, pk):
    obj = get_object_or_404(Specialization, pk=pk)
    if request.method == 'POST':
        if obj.doctors.exists():
            obj.status = 'Inactive'
            obj.save(update_fields=['status'])
            messages.warning(request, f'Specialization has linked doctors – deactivated instead of deleted.')
        else:
            name = obj.name
            obj.delete()
            messages.success(request, f'Specialization "{name}" deleted.')
            log_action(request.user, 'delete', 'masters', pk, f'Specialization {name}', request=request)
    return redirect('masters:specialization_list')


# ═══════════════════════════════════════════════════════════
# HOD
# ═══════════════════════════════════════════════════════════
@admin_required
def hod_list(request):
    q = request.GET.get('q', '').strip()
    qs = HOD.objects.select_related('department').order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(employee_id__icontains=q) | Q(email__icontains=q))
    return render(request, 'masters/hod_list.html', {
        'page_obj': _paginate(request, qs), 'q': q,
    })


@admin_required
def hod_create(request):
    form = HODForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        log_action(request.user, 'create', 'masters', obj.pk, f'HOD {obj}', request=request)
        messages.success(request, f'HOD "{obj.name}" created for {obj.department}.')
        return redirect('masters:hod_list')
    return render(request, 'masters/hod_form.html', {'form': form, 'title': 'Add HOD'})


@admin_required
def hod_edit(request, pk):
    obj = get_object_or_404(HOD, pk=pk)
    form = HODForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f'HOD "{obj.name}" updated.')
        return redirect('masters:hod_list')
    return render(request, 'masters/hod_form.html', {'form': form, 'title': f'Edit HOD – {obj.name}', 'object': obj})


@admin_required
def hod_deactivate(request, pk):
    obj = get_object_or_404(HOD, pk=pk)
    if request.method == 'POST':
        obj.status = 'Inactive' if obj.status == 'Active' else 'Active'
        obj.save(update_fields=['status'])
        messages.success(request, f'HOD "{obj.name}" set to {obj.status}.')
    return redirect('masters:hod_list')


# ═══════════════════════════════════════════════════════════
# DOCTOR MASTER
# ═══════════════════════════════════════════════════════════
@admin_required
def doctor_list(request):
    q = request.GET.get('q', '').strip()
    spec = request.GET.get('specialization', '')
    status = request.GET.get('status', '')
    qs = Doctor.objects.select_related('specialization', 'department').order_by('name')
    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(doctor_id__icontains=q) |
            Q(registration_number__icontains=q) | Q(mobile__icontains=q) | Q(email__icontains=q)
        )
    if spec:
        qs = qs.filter(specialization_id=spec)
    if status:
        qs = qs.filter(status=status)
    return render(request, 'masters/doctor_list.html', {
        'page_obj': _paginate(request, qs), 'q': q, 'spec': spec, 'status': status,
        'specializations': Specialization.objects.filter(status='Active'),
    })


@admin_required
def doctor_detail(request, pk):
    doctor = get_object_or_404(
        Doctor.objects.select_related('specialization', 'department'), pk=pk
    )
    schedules = doctor.schedules.filter(is_active=True).order_by('day_of_week', 'start_time')
    return render(request, 'masters/doctor_detail.html', {'doctor': doctor, 'schedules': schedules})


@admin_required
def doctor_create(request):
    form = DoctorForm(request.POST or None, request.FILES or None)
    formset = DoctorScheduleFormSet(request.POST or None)
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            doctor = form.save()
            formset.instance = doctor
            formset.save()
        log_action(request.user, 'create', 'doctors', doctor.pk, f'Doctor {doctor}', request=request)
        messages.success(request, f'Doctor "{doctor.name}" added successfully.')
        return redirect('masters:doctor_list')
    return render(request, 'masters/doctor_form.html', {
        'form': form, 'formset': formset, 'title': 'Add Doctor',
    })


@admin_required
def doctor_edit(request, pk):
    doctor = get_object_or_404(Doctor, pk=pk)
    form = DoctorForm(request.POST or None, request.FILES or None, instance=doctor)
    formset = DoctorScheduleFormSet(request.POST or None, instance=doctor)
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            form.save()
            formset.save()
        log_action(request.user, 'update', 'doctors', doctor.pk, f'Doctor {doctor}', request=request)
        messages.success(request, f'Doctor "{doctor.name}" updated.')
        return redirect('masters:doctor_detail', pk=doctor.pk)
    return render(request, 'masters/doctor_form.html', {
        'form': form, 'formset': formset, 'title': f'Edit Doctor – {doctor.name}', 'object': doctor,
    })


@admin_required
def doctor_deactivate(request, pk):
    doctor = get_object_or_404(Doctor, pk=pk)
    if request.method == 'POST':
        doctor.status = 'Inactive' if doctor.status == 'Active' else 'Active'
        doctor.save(update_fields=['status'])
        log_action(request.user, 'status_change', 'doctors', doctor.pk, f'{doctor} → {doctor.status}', request=request)
        messages.success(request, f'Doctor "{doctor.name}" set to {doctor.status}.')
    return redirect('masters:doctor_list')


# ═══════════════════════════════════════════════════════════
# STUDENT MASTER
# ═══════════════════════════════════════════════════════════
@admin_required
def student_list(request):
    q = request.GET.get('q', '').strip()
    dept = request.GET.get('department', '')
    status = request.GET.get('status', '')
    qs = Patient.objects.filter(category=Patient.CATEGORY_STUDENT).select_related(
        'student_profile', 'student_profile__programme', 'student_profile__department'
    ).order_by('name')
    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(patient_id__icontains=q) | Q(mobile__icontains=q) |
            Q(email__icontains=q) | Q(student_profile__enrollment_number__icontains=q)
        )
    if dept:
        qs = qs.filter(student_profile__department_id=dept)
    if status:
        qs = qs.filter(status=status)
    return render(request, 'masters/student_list.html', {
        'page_obj': _paginate(request, qs), 'q': q, 'dept': dept, 'status': status,
        'departments': Department.objects.filter(status='Active'),
    })


@admin_required
def student_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STUDENT)
    return render(request, 'masters/student_detail.html', {'patient': patient})


@admin_required
def student_create(request):
    form = StudentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            # Duplicate checks
            pid = form.cleaned_data['patient_id']
            enr = form.cleaned_data['enrollment_number']
            if Patient.objects.filter(patient_id=pid).exists():
                form.add_error('patient_id', 'A patient with this ID already exists.')
            elif StudentProfile.objects.filter(enrollment_number=enr).exists():
                form.add_error('enrollment_number', 'Enrollment number already exists.')
            else:
                patient = form.save(commit=False)
                patient.category = Patient.CATEGORY_STUDENT
                patient.created_by = request.user
                patient.save()
                StudentProfile.objects.create(
                    patient=patient,
                    enrollment_number=enr,
                    programme=form.cleaned_data.get('programme'),
                    department=form.cleaned_data.get('department'),
                    semester=form.cleaned_data.get('semester'),
                )
                log_action(request.user, 'create', 'patients', patient.pk, f'Student {patient}', request=request)
                messages.success(request, f'Student "{patient.name}" created.')
                return redirect('masters:student_list')
    return render(request, 'masters/student_form.html', {'form': form, 'title': 'Add Student'})


@admin_required
def student_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STUDENT)
    form = StudentForm(request.POST or None, instance=patient)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            enr = form.cleaned_data['enrollment_number']
            if StudentProfile.objects.filter(enrollment_number=enr).exclude(patient=patient).exists():
                form.add_error('enrollment_number', 'Enrollment number already exists.')
            else:
                patient = form.save()
                sp, _ = StudentProfile.objects.get_or_create(patient=patient)
                sp.enrollment_number = enr
                sp.programme = form.cleaned_data.get('programme')
                sp.department = form.cleaned_data.get('department')
                sp.semester = form.cleaned_data.get('semester')
                sp.save()
                log_action(request.user, 'update', 'patients', patient.pk, f'Student {patient}', request=request)
                messages.success(request, f'Student "{patient.name}" updated.')
                return redirect('masters:student_detail', pk=patient.pk)
    return render(request, 'masters/student_form.html', {
        'form': form, 'title': f'Edit Student – {patient.name}', 'object': patient,
    })


@admin_required
def student_deactivate(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STUDENT)
    if request.method == 'POST':
        patient.status = 'Inactive' if patient.status == 'Active' else 'Active'
        patient.save(update_fields=['status'])
        messages.success(request, f'Student "{patient.name}" set to {patient.status}.')
    return redirect('masters:student_list')


@admin_required
def student_export(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'patient_id', 'enrollment_no', 'name', 'gender', 'dob', 'programme',
        'department', 'semester', 'email', 'mobile', 'blood_group', 'status',
    ])
    qs = Patient.objects.filter(category=Patient.CATEGORY_STUDENT).select_related(
        'student_profile', 'student_profile__programme', 'student_profile__department'
    )
    for p in qs:
        sp = getattr(p, 'student_profile', None)
        writer.writerow([
            p.patient_id,
            sp.enrollment_number if sp else '',
            p.name, p.gender, p.date_of_birth or '',
            sp.programme.name if sp and sp.programme else '',
            sp.department.name if sp and sp.department else '',
            sp.semester if sp else '',
            p.email, p.mobile, p.blood_group, p.status,
        ])
    log_action(request.user, 'export', 'patients', description='Student export CSV', request=request)
    return response


@admin_required
def student_import(request):
    form = StudentImportForm(request.POST or None, request.FILES or None)
    results = None
    if request.method == 'POST' and form.is_valid():
        f = form.cleaned_data['file']
        name = f.name.lower()
        rows = []
        try:
            if name.endswith('.csv'):
                text = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
            elif name.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                messages.error(request, 'Unsupported file type. Use CSV or Excel.')
                return render(request, 'masters/student_import.html', {'form': form})
        except Exception as e:
            messages.error(request, f'Could not read file: {e}')
            return render(request, 'masters/student_import.html', {'form': form})

        success, failed = [], []
        with transaction.atomic():
            for i, row in enumerate(rows, start=2):
                try:
                    # Normalize keys
                    r = {str(k).strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
                    pid = str(r.get('patient_id') or r.get('student_id') or '').strip()
                    enr = str(r.get('enrollment_no') or r.get('enrollment_number') or '').strip()
                    name_val = str(r.get('name') or '').strip()
                    if not pid or not enr or not name_val:
                        failed.append({'row': i, 'reason': 'Missing patient_id / enrollment_no / name'})
                        continue
                    if Patient.objects.filter(patient_id=pid).exists():
                        failed.append({'row': i, 'reason': f'Duplicate patient_id {pid}'})
                        continue
                    if StudentProfile.objects.filter(enrollment_number=enr).exists():
                        failed.append({'row': i, 'reason': f'Duplicate enrollment {enr}'})
                        continue
                    dept = None
                    prog = None
                    dname = str(r.get('department') or '').strip()
                    pname = str(r.get('programme') or '').strip()
                    if dname:
                        dept = Department.objects.filter(Q(name__iexact=dname) | Q(department_code__iexact=dname)).first()
                    if pname:
                        prog = Programme.objects.filter(Q(name__iexact=pname) | Q(programme_code__iexact=pname)).first()
                    patient = Patient.objects.create(
                        patient_id=pid, category=Patient.CATEGORY_STUDENT, name=name_val,
                        gender=str(r.get('gender') or '').strip()[:10],
                        email=str(r.get('email') or '').strip(),
                        mobile=str(r.get('mobile') or '').strip()[:15],
                        blood_group=str(r.get('blood_group') or '').strip()[:5],
                        status='Active', created_by=request.user,
                    )
                    StudentProfile.objects.create(
                        patient=patient, enrollment_number=enr,
                        programme=prog, department=dept,
                        semester=int(r['semester']) if r.get('semester') not in (None, '') else None,
                    )
                    success.append(pid)
                except Exception as e:
                    failed.append({'row': i, 'reason': str(e)})
        results = {
            'total': len(rows), 'success': len(success), 'failed': len(failed),
            'failed_rows': failed[:50],
        }
        log_action(
            request.user, 'bulk_import', 'patients',
            description=f'Student import: {len(success)} ok, {len(failed)} failed',
            request=request,
        )
        messages.success(request, f'Import finished: {len(success)} imported, {len(failed)} failed.')
    return render(request, 'masters/student_import.html', {'form': form, 'results': results})


@admin_required
def student_import_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'patient_id', 'enrollment_no', 'name', 'gender', 'dob', 'programme',
        'department', 'semester', 'email', 'mobile', 'blood_group',
    ])
    writer.writerow([
        'P-STU-9001', 'GV20269001', 'Sample Student', 'Male', '2003-05-15',
        'Bachelor of Computer Applications', 'Computer Science', '3',
        'sample@gujaratvidyapith.ac.in', '9876543210', 'B+',
    ])
    return response


# ═══════════════════════════════════════════════════════════
# STAFF MASTER
# ═══════════════════════════════════════════════════════════
@admin_required
def staff_list(request):
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    qs = Patient.objects.filter(category=Patient.CATEGORY_STAFF).select_related(
        'staff_profile', 'staff_profile__department'
    ).order_by('name')
    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(patient_id__icontains=q) | Q(mobile__icontains=q) |
            Q(staff_profile__employee_id__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)
    return render(request, 'masters/staff_list.html', {
        'page_obj': _paginate(request, qs), 'q': q, 'status': status,
    })


@admin_required
def staff_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STAFF)
    family = []
    if hasattr(patient, 'staff_profile'):
        family = StaffFamilyProfile.objects.filter(related_staff=patient.staff_profile).select_related('patient')
    return render(request, 'masters/staff_detail.html', {'patient': patient, 'family': family})


@admin_required
def staff_create(request):
    form = StaffForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            pid = form.cleaned_data['patient_id']
            eid = form.cleaned_data['employee_id']
            if Patient.objects.filter(patient_id=pid).exists():
                form.add_error('patient_id', 'Patient ID already exists.')
            elif StaffProfile.objects.filter(employee_id=eid).exists():
                form.add_error('employee_id', 'Employee ID already exists.')
            else:
                patient = form.save(commit=False)
                patient.category = Patient.CATEGORY_STAFF
                patient.created_by = request.user
                patient.save()
                StaffProfile.objects.create(
                    patient=patient, employee_id=eid,
                    department=form.cleaned_data.get('department'),
                    designation=form.cleaned_data.get('designation') or '',
                )
                messages.success(request, f'Staff "{patient.name}" created.')
                return redirect('masters:staff_list')
    return render(request, 'masters/staff_form.html', {'form': form, 'title': 'Add Staff'})


@admin_required
def staff_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STAFF)
    form = StaffForm(request.POST or None, instance=patient)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            eid = form.cleaned_data['employee_id']
            if StaffProfile.objects.filter(employee_id=eid).exclude(patient=patient).exists():
                form.add_error('employee_id', 'Employee ID already exists.')
            else:
                patient = form.save()
                sp, _ = StaffProfile.objects.get_or_create(patient=patient)
                sp.employee_id = eid
                sp.department = form.cleaned_data.get('department')
                sp.designation = form.cleaned_data.get('designation') or ''
                sp.save()
                messages.success(request, f'Staff "{patient.name}" updated.')
                return redirect('masters:staff_detail', pk=patient.pk)
    return render(request, 'masters/staff_form.html', {
        'form': form, 'title': f'Edit Staff – {patient.name}', 'object': patient,
    })


@admin_required
def staff_deactivate(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STAFF)
    if request.method == 'POST':
        patient.status = 'Inactive' if patient.status == 'Active' else 'Active'
        patient.save(update_fields=['status'])
        messages.success(request, f'Staff "{patient.name}" set to {patient.status}.')
    return redirect('masters:staff_list')


# ═══════════════════════════════════════════════════════════
# STAFF FAMILY MASTER
# ═══════════════════════════════════════════════════════════
@admin_required
def family_list(request):
    q = request.GET.get('q', '').strip()
    qs = Patient.objects.filter(category=Patient.CATEGORY_STAFF_FAMILY).select_related(
        'family_profile', 'family_profile__related_staff', 'family_profile__related_staff__patient'
    ).order_by('name')
    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(patient_id__icontains=q) | Q(mobile__icontains=q) |
            Q(family_profile__related_staff__patient__name__icontains=q)
        )
    return render(request, 'masters/family_list.html', {
        'page_obj': _paginate(request, qs), 'q': q,
    })


@admin_required
def family_create(request):
    form = StaffFamilyForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            pid = form.cleaned_data['patient_id']
            if Patient.objects.filter(patient_id=pid).exists():
                form.add_error('patient_id', 'Patient ID already exists.')
            else:
                patient = form.save(commit=False)
                patient.category = Patient.CATEGORY_STAFF_FAMILY
                patient.created_by = request.user
                patient.save()
                StaffFamilyProfile.objects.create(
                    patient=patient,
                    related_staff=form.cleaned_data['related_staff'],
                    relationship=form.cleaned_data['relationship'],
                )
                messages.success(request, f'Family member "{patient.name}" created.')
                return redirect('masters:family_list')
    return render(request, 'masters/family_form.html', {'form': form, 'title': 'Add Staff Family Member'})


@admin_required
def family_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STAFF_FAMILY)
    form = StaffFamilyForm(request.POST or None, instance=patient)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            patient = form.save()
            fp, _ = StaffFamilyProfile.objects.get_or_create(patient=patient)
            fp.related_staff = form.cleaned_data['related_staff']
            fp.relationship = form.cleaned_data['relationship']
            fp.save()
            messages.success(request, f'Family member "{patient.name}" updated.')
            return redirect('masters:family_list')
    return render(request, 'masters/family_form.html', {
        'form': form, 'title': f'Edit Family Member – {patient.name}', 'object': patient,
    })


@admin_required
def family_deactivate(request, pk):
    patient = get_object_or_404(Patient, pk=pk, category=Patient.CATEGORY_STAFF_FAMILY)
    if request.method == 'POST':
        patient.status = 'Inactive' if patient.status == 'Active' else 'Active'
        patient.save(update_fields=['status'])
        messages.success(request, f'Family member "{patient.name}" set to {patient.status}.')
    return redirect('masters:family_list')
