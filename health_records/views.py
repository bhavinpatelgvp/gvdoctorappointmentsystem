import csv
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone

from accounts.permissions import doctor_required, admin_required
from accounts.authorization import require_patient_owner, patient_scoped_qs, get_linked_patient, get_doctor_profile
from audit.services import log_action
from patients.models import Patient
from doctors.models import Doctor
from .models import (
    HealthCheckup, CBCReport, RBSReport, BloodPressureReport,
    LipidProfileReport, ClinicalParameter, ClinicalParameterValue, BulkImportLog,
)
from .forms import HealthCheckupForm, CBCForm, RBSForm, BPForm, LipidForm, OtherParamsForm, BulkHealthImportForm


@login_required
def index(request):
    q = request.GET.get('q', '').strip()
    qs = HealthCheckup.objects.select_related('patient', 'doctor').order_by('-checkup_date')
    qs = patient_scoped_qs(request.user, qs)
    if q:
        from django.db.models import Q
        qs = qs.filter(Q(patient__name__icontains=q) | Q(patient__patient_id__icontains=q))
    page = Paginator(qs, 10).get_page(request.GET.get('page'))
    return render(request, 'health_records/list.html', {
        'page_obj': page, 'checkups': page, 'q': q,
    })


@doctor_required
def entry_start(request):
    """Patient search → start health check-up."""
    q = request.GET.get('q', '').strip()
    patients = []
    if q:
        from django.db.models import Q
        patients = Patient.objects.filter(status='Active').filter(
            Q(patient_id__icontains=q) | Q(name__icontains=q) | Q(mobile__icontains=q)
        )[:20]
    return render(request, 'health_records/entry_start.html', {'q': q, 'patients': patients})


@doctor_required
def entry(request, patient_id):
    patient = get_object_or_404(Patient, patient_id=patient_id)
    doctor = get_doctor_profile(request.user)
    checkup_form = HealthCheckupForm(request.POST or None, initial={'checkup_date': date.today()})
    cbc_form = CBCForm(request.POST or None, prefix='cbc')
    rbs_form = RBSForm(request.POST or None, prefix='rbs')
    bp_form = BPForm(request.POST or None, prefix='bp', initial={'measured_at': timezone.now()})
    lipid_form = LipidForm(request.POST or None, prefix='lipid')
    other_params = ClinicalParameter.objects.filter(category='other', is_active=True)
    other_form = OtherParamsForm(request.POST or None, prefix='other', parameters=other_params)

    if request.method == 'POST' and checkup_form.is_valid():
        with transaction.atomic():
            hc = checkup_form.save(commit=False)
            hc.patient = patient
            hc.doctor = doctor
            hc.entered_by = request.user
            hc.save()
            # CBC
            if any(request.POST.get(f'cbc-{f}') for f in CBCForm.Meta.fields if f != 'remarks'):
                if cbc_form.is_valid():
                    cbc = cbc_form.save(commit=False)
                    cbc.health_checkup = hc
                    cbc.entered_by = request.user
                    cbc.save()
                    # Mirror HB to parameter values for analytics
                    _mirror_param(hc, 'HB', cbc.hemoglobin, patient.gender)
            # RBS
            if request.POST.get('rbs-value'):
                if rbs_form.is_valid():
                    rbs = rbs_form.save(commit=False)
                    rbs.health_checkup = hc
                    rbs.entered_by = request.user
                    rbs.save()
                    _mirror_param(hc, 'RBS', rbs.value, patient.gender)
            # BP
            if request.POST.get('bp-systolic'):
                if bp_form.is_valid():
                    bp = bp_form.save(commit=False)
                    bp.health_checkup = hc
                    bp.recorded_by = request.user
                    if not bp.measured_at:
                        bp.measured_at = timezone.now()
                    bp.save()
                    _mirror_param(hc, 'SYS', bp.systolic, patient.gender)
                    _mirror_param(hc, 'DIA', bp.diastolic, patient.gender)
            # Lipid
            if any(request.POST.get(f'lipid-{f}') for f in ['total_cholesterol', 'hdl', 'ldl', 'triglycerides']):
                if lipid_form.is_valid():
                    lip = lipid_form.save(commit=False)
                    lip.health_checkup = hc
                    lip.entered_by = request.user
                    lip.save()
                    _mirror_param(hc, 'TC', lip.total_cholesterol, patient.gender)
            # Other
            if other_form.is_valid():
                for code, val in other_form.cleaned_data.items():
                    if val is not None:
                        _mirror_param(hc, code, val, patient.gender)
        log_action(request.user, 'create', 'health_records', hc.pk, f'Checkup {patient}', request=request)
        messages.success(request, 'Health check-up saved.')
        return redirect('health_records:detail', pk=hc.pk)

    return render(request, 'health_records/entry.html', {
        'patient': patient,
        'checkup_form': checkup_form,
        'cbc_form': cbc_form,
        'rbs_form': rbs_form,
        'bp_form': bp_form,
        'lipid_form': lipid_form,
        'other_form': other_form,
    })


def _mirror_param(hc, code, value, gender):
    if value is None:
        return
    param = ClinicalParameter.objects.filter(code=code, is_active=True).first()
    if not param:
        return
    cpv, _ = ClinicalParameterValue.objects.update_or_create(
        health_checkup=hc, parameter=param,
        defaults={'value': value, 'unit': param.unit},
    )
    cpv.evaluate_status(gender)
    cpv.save(update_fields=['status'])


@login_required
def detail(request, pk):
    hc = get_object_or_404(
        HealthCheckup.objects.select_related('patient', 'doctor').prefetch_related(
            'parameter_values__parameter', 'bp_reports',
        ),
        pk=pk,
    )
    require_patient_owner(request.user, hc.patient)
    cbc = getattr(hc, 'cbc_report', None)
    rbs = getattr(hc, 'rbs_report', None)
    lipid = getattr(hc, 'lipid_report', None)
    return render(request, 'health_records/detail.html', {
        'checkup': hc, 'cbc': cbc, 'rbs': rbs, 'lipid': lipid,
        'bp_list': hc.bp_reports.all(),
        'params': hc.parameter_values.select_related('parameter').all(),
    })


@doctor_required
def bulk_import(request):
    form = BulkHealthImportForm(request.POST or None, request.FILES or None)
    results = None
    if request.method == 'POST' and form.is_valid():
        f = form.cleaned_data['file']
        name = f.name.lower()
        rows = []
        try:
            if name.endswith('.csv'):
                text = f.read().decode('utf-8-sig')
                rows = list(csv.DictReader(io.StringIO(text)))
            else:
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [v for v in row])))
        except Exception as e:
            messages.error(request, f'Could not read file: {e}')
            return render(request, 'health_records/bulk_import.html', {'form': form})

        success, failed = 0, []
        doctor = get_doctor_profile(request.user)
        with transaction.atomic():
            for i, row in enumerate(rows, start=2):
                try:
                    r = {str(k).strip().lower(): v for k, v in row.items() if k}
                    pid = str(r.get('patient_id') or '').strip()
                    if not pid:
                        failed.append({'row': i, 'reason': 'Missing patient_id'})
                        continue
                    patient = Patient.objects.filter(patient_id=pid).first()
                    if not patient:
                        failed.append({'row': i, 'reason': f'Patient not found: {pid}'})
                        continue
                    td = r.get('test_date')
                    if isinstance(td, datetime):
                        tdate = td.date()
                    elif isinstance(td, date):
                        tdate = td
                    else:
                        tdate = datetime.strptime(str(td).strip()[:10], '%Y-%m-%d').date() if td else date.today()
                    hc = HealthCheckup.objects.create(
                        patient=patient, doctor=doctor, checkup_date=tdate,
                        entered_by=request.user, notes='Bulk import', is_demo=False,
                    )
                    def num(key):
                        v = r.get(key)
                        if v in (None, ''):
                            return None
                        return Decimal(str(v))
                    hb = num('hb') or num('hemoglobin')
                    if hb is not None or num('wbc') or num('rbc'):
                        CBCReport.objects.create(
                            health_checkup=hc, hemoglobin=hb,
                            rbc_count=num('rbc'), wbc_count=num('wbc'),
                            platelet_count=num('platelets'), hematocrit=num('pcv') or num('hematocrit'),
                            mcv=num('mcv'), mch=num('mch'), mchc=num('mchc'),
                            neutrophils=num('neutrophils'), lymphocytes=num('lymphocytes'),
                            monocytes=num('monocytes'), eosinophils=num('eosinophils'),
                            basophils=num('basophils'), entered_by=request.user,
                        )
                        _mirror_param(hc, 'HB', hb, patient.gender)
                    rbs_v = num('rbs')
                    if rbs_v is not None:
                        RBSReport.objects.create(health_checkup=hc, value=rbs_v, entered_by=request.user)
                        _mirror_param(hc, 'RBS', rbs_v, patient.gender)
                    sys_v, dia_v = num('systolic_bp') or num('systolic'), num('diastolic_bp') or num('diastolic')
                    if sys_v is not None and dia_v is not None:
                        BloodPressureReport.objects.create(
                            health_checkup=hc, systolic=int(sys_v), diastolic=int(dia_v),
                            pulse_rate=int(num('pulse')) if num('pulse') else None,
                            measured_at=timezone.make_aware(datetime.combine(tdate, datetime.min.time())),
                            recorded_by=request.user,
                        )
                    tc = num('total_cholesterol') or num('cholesterol')
                    if tc is not None:
                        LipidProfileReport.objects.create(
                            health_checkup=hc, total_cholesterol=tc,
                            hdl=num('hdl'), ldl=num('ldl'), triglycerides=num('triglycerides'),
                            entered_by=request.user,
                        )
                    for code in ('height', 'weight', 'bmi'):
                        v = num(code) or num(code.upper()) if False else num(code)
                        # map keys
                        v = num('height') if code == 'height' else (num('weight') if code == 'weight' else num('bmi'))
                        if v is not None:
                            _mirror_param(hc, {'height': 'HT', 'weight': 'WT', 'bmi': 'BMI'}[code], v, patient.gender)
                    success += 1
                except Exception as e:
                    failed.append({'row': i, 'reason': str(e)})
            BulkImportLog.objects.create(
                import_type='health_checkup',
                file_name=f.name,
                total_records=len(rows),
                success_count=success,
                failed_count=len(failed),
                status='completed' if not failed else ('partial' if success else 'failed'),
                summary=f'{success} ok, {len(failed)} failed',
                imported_by=request.user,
            )
        results = {'total': len(rows), 'success': success, 'failed': len(failed), 'failed_rows': failed[:100]}
        log_action(request.user, 'bulk_import', 'health_records', description=f'{success}/{len(rows)}', request=request)
        messages.info(request, f'Import: {success} success, {len(failed)} failed.')
    return render(request, 'health_records/bulk_import.html', {'form': form, 'results': results})


@doctor_required
def bulk_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="health_checkup_import_template.csv"'
    w = csv.writer(response)
    w.writerow([
        'patient_id', 'test_date', 'hb', 'rbc', 'wbc', 'platelets', 'pcv', 'mcv', 'mch', 'mchc',
        'neutrophils', 'lymphocytes', 'monocytes', 'eosinophils', 'basophils',
        'rbs', 'systolic_bp', 'diastolic_bp', 'pulse',
        'total_cholesterol', 'hdl', 'ldl', 'triglycerides', 'height', 'weight', 'bmi',
    ])
    w.writerow([
        'P-STU-0001', '2026-08-01', '13.2', '4.8', '7500', '250000', '40', '88', '29', '33',
        '55', '32', '6', '3', '1', '110', '120', '80', '72', '190', '45', '110', '140', '170', '65', '22.5',
    ])
    return response
